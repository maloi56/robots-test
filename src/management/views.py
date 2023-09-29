from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.db.models import Count
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import TemplateView
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook

from common.view import TitleMixin
from robots.models import Robot
from users.models import User

from itertools import groupby
from operator import itemgetter


class ManagementView(TitleMixin, LoginRequiredMixin, TemplateView):
    template_name = 'management/management.html'
    login_url = reverse_lazy('users:login')
    title = 'R4C - management'
    model = Robot
    context_object_name = 'result'

    def get(self, request, *args, **kwargs):
        if request.user.role != User.DIRECTOR:
            raise PermissionDenied
        return super().get(request, *args, **kwargs)


@login_required
def get_excel_report(request, days_count):
    """
        Получение отчета произведенных роботов за :param days_count: дней
    """

    if request.user.role != User.DIRECTOR:
        raise PermissionDenied

    workbook = Workbook()
    workbook.remove(workbook.active)
    start_date = timezone.now() - timezone.timedelta(days=days_count)

    queryset = (
        Robot.objects.filter(created__gte=start_date)
        .values('model', 'version')
        .annotate(count=Count('pk'))
        .order_by('-count')
    )

    queryset = sorted(queryset, key=itemgetter('model', 'version'))

    for model, group in groupby(queryset, key=itemgetter('model')):
        sheet = workbook.create_sheet(title=f'{model}')
        sheet.append(('Модель', 'Версия', f'Количество за {days_count} дней'))

        for version in group:
            sheet.append((version['model'], version['version'], version['count']))

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

        sheet.column_dimensions['C'].width = 2 * sheet.column_dimensions['C'].width

    now = timezone.now().date().strftime('%d-%m')
    start_date_str = start_date.date().strftime('%d-%m')
    filename = f'R4C_report_for_{start_date_str}:{now}_({days_count} days).xlsx'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    workbook.save(response)

    return response
