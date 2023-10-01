from itertools import groupby
from operator import itemgetter

from django.db.models import Count
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook

from robots.models import Robot


def generate_created_robots_per_period_report(days_count, start_date):
    workbook = Workbook()
    workbook.remove(workbook.active)

    queryset = (
        Robot.objects.filter(created__gte=start_date)
        .values('model', 'version')
        .annotate(count=Count('pk'))
        .order_by('-count')
    )

    queryset = sorted(queryset, key=itemgetter('model', 'version'))

    for model, group in groupby(queryset, key=itemgetter('model')):
        sheet = workbook.create_sheet(title=f'{model}')
        sheet.append(('Модель', 'Версия', f'Количество за {days_count} дней'))

        for version in group:
            sheet.append((version['model'], version['version'], version['count']))

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

        sheet.column_dimensions['C'].width = 2 * sheet.column_dimensions['C'].width

    return workbook
